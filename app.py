import sys
sys.path.append('./libs')

from flask import Flask, render_template, request
import os
import glob
from openpyxl import load_workbook
import pyperclip
import re

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        input_text = request.form['input_text']
        row_number = int(request.form['row_number'])
        excel_file = request.files['excel_file']
        
        # Save the uploaded file
        excel_file.save(excel_file.filename)
        selected_file = check_file_type(excel_file.filename)
        
        if selected_file:
            replaced_text = replace_special_strings(input_text, row_number, selected_file)
            pyperclip.copy(replaced_text)  # Copy replaced text to clipboard
            return render_template('index.html', replaced_text=replaced_text)
        else:
            return "Please select an Excel file."
    return render_template('index.html', replaced_text=None)

def check_file_type(file_path):
    if file_path.lower().endswith('.numbers'):
        return None
    return file_path

def replace_special_strings(text, row_number, excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    
    def replace_match(match):
        col_label = match.group(1)
        col_index = ord(col_label) - ord('A') + 1  # openpyxl column index starts from 1
        cell_value = ws.cell(row=row_number, column=col_index).value
        return str(cell_value) if cell_value is not None else ''
    
    pattern = re.compile(r'&([A-Z]+)&')
    replaced_text = pattern.sub(replace_match, text)
    return replaced_text

if __name__ == '__main__':
    app.run(debug=True)
